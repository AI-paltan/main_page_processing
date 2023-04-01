import openpyxl
import pandas as pd
import numpy as np
from .database import get_db1
import db_models
from .data_dump_config import datadump_core_settings

db =get_db1()


class DataDump:
    def __init__(self,fileid,meta_dict,cbs_resposne_bucket,cpl_resposne_bucket,ccf_resposne_bucket) -> None:
        self.fileid = fileid
        self.years_list = None
        self.cbs_resposne_bucket = cbs_resposne_bucket
        self.cpl_response_bucket = cpl_resposne_bucket
        self.ccf_response_bucket = ccf_resposne_bucket
        self.workbook = None
        self.month = None
        self.meta_dict = meta_dict

    def trigger_job(self):
        self.load_workbook()
        self.get_year_list()
        self.set_years_header()
        self.dump_cbs_data()
        self.dump_cpl_data()
        self.dump_ccf_data()
        self.save_excel()

    def get_year_list(self):
        for k,v in self.meta_dict.items():
            # print(v)
            self.years_list = v["year_list"]
            self.years_list = sorted(self.years_list)
            break

    def get_crm_nlp_df(self,statement_type):
        df_crm_nlp_bucket_master = db.query(db_models.CRM_nlp_bucketing)
        crm_nlp_df = pd.read_sql(df_crm_nlp_bucket_master.statement, df_crm_nlp_bucket_master.session.bind)
        bs_crm_nlp_df = crm_nlp_df[crm_nlp_df['statement_type']==statement_type]
        bs_crm_nlp_df['sort'] = pd.to_numeric(bs_crm_nlp_df['cdm_keyword_start_row_map'],errors="coerce")
        bs_crm_nlp_df.sort_values('sort',inplace=True, ascending=False)
        bs_crm_nlp_df = bs_crm_nlp_df.drop('sort', axis=1)
        # bs_crm_nlp_df_sorted_rev = bs_crm_nlp_df.sort_values(by='cdm_keyword_start_row_map',ascending=False)
        return bs_crm_nlp_df
    
    def get_month(self,):
        file_query = db.query(db_models.FileLogs).filter(db_models.FileLogs.fileid == self.fileid).first()
        self.month = file_query.month
    
    def set_years_header(self):
        max_year = max(self.years_list)
        financial_year_pattern = str(max_year)
        if self.month == "december":
            financial_year_pattern = str(max_year)+ '/' + str('12')
        if self.month == "march":
            financial_year_pattern = str(max_year)+ '/' + str('03')
        worksheet = self.workbook.get_sheet_by_name('BS')
        worksheet.cell(row=11, column=2).value = financial_year_pattern

    def get_years_excel_colmap(self,total_col):
        year_colidx = total_col - len(self.years_list)
        year_excel_col_map_dict = {}
        for i in range(len(self.years_list)):
            col_no = year_colidx + i
            year_excel_col_map_dict[self.years_list[i]] = col_no
        return year_excel_col_map_dict
    
    def insert_rows(self,worksheet,df_len,start_template_rwo,total_end_template_row):
        if df_len >= abs(int(total_end_template_row)-int(start_template_rwo)):
            row_amount = int((df_len - abs(int(total_end_template_row)-int(start_template_rwo))) +3)
            # print(start_template_rwo,row_amount)
            start_idx = int(start_template_rwo)+1
            worksheet.insert_rows(start_idx,row_amount)
        return worksheet
    
    def insert_records(self,bs_crm_nlp_df_sorted,cbs_resposne_bucket,years_list,year_excel_col_map_dict):
    # print(cbs_resposne_bucket)
        for idx,row in bs_crm_nlp_df_sorted.iterrows():
            # print(idx)
            # print(row["meta_keyword"])
            if row["cdm_keyword_start_row_map"] is not None:
                if row["cdm_keyword_start_row_map"].isdigit():
                    cbs_worksheet = self.report_workbook.get_sheet_by_name(row['cdm_sheet_name'])
                    repsonse_dict = cbs_resposne_bucket.get(row['meta_keyword'])
                    notes_horizontal_table_df = repsonse_dict.get('notes_horizontal_table_df')  
                    # total_row_num  = int(row["cdm_total_row_map"]) 
                    # cbs_worksheet.cell(row=total_row_num,column=year_column).value = row_note[year]
                    if len(notes_horizontal_table_df)>0:
                        # print("len: ", len(notes_horizontal_table_df))
                        cbs_worksheet = self.insert_rows(worksheet=cbs_worksheet,df_len=len(notes_horizontal_table_df),start_template_rwo=row["cdm_keyword_start_row_map"],total_end_template_row=row["cdm_total_row_map"])
                        for idx,row_note in notes_horizontal_table_df.iterrows():
                            excel_row_num = int(row["cdm_keyword_start_row_map"]) + idx+1
                            cbs_worksheet.cell(row=excel_row_num,column=2).value = row_note["line_item"]
                            for year in years_list:
                                year_column = year_excel_col_map_dict.get(int(year))
                                try:
                                    cbs_worksheet.cell(row=excel_row_num,column=year_column).value = row_note[year]
                                except:
                                    cbs_worksheet.cell(row=excel_row_num,column=year_column).value = 0.0
                    total_row_num  = int(row["cdm_total_row_map"]) 
                    for main_year,value in zip(repsonse_dict.get("main_page_year_list"),repsonse_dict.get("main_page_year_total")):
                        year_column = year_excel_col_map_dict.get(int(year))
                        try:
                            cbs_worksheet.cell(row=total_row_num,column=year_column).value = value
                        except:
                            cbs_worksheet.cell(row=total_row_num,column=year_column).value = 0.0


    def get_row_map_from_db(self,meta_keyword):
        meta_query = db.query(db_models.CRM_nlp_bucketing).filter(db_models.CRM_nlp_bucketing.meta_keyword == meta_keyword).first()
        cdm_sheet_name = meta_query.cdm_sheet_name
        cdm_keyword_start_row_map = meta_query.cdm_keyword_start_row_map
        cdm_total_row_map = meta_query.cdm_total_row_map
        return cdm_sheet_name,cdm_keyword_start_row_map,cdm_total_row_map
    
    def load_workbook(self):
        self.workbook = openpyxl.load_workbook(datadump_core_settings.cdm_template)

    def set_year_to_excel(self):
        pass

    def dump_cbs_data(self):
        bs_crm_nlp_df_sorted_rev = self.get_crm_nlp_df(statement_type="cbs")
        # month = get_month(fileid)
        year_excel_col_map_dict = self.get_years_excel_colmap(total_col=9)
        self.insert_records(report_workbook=self.workbook,bs_crm_nlp_df_sorted=bs_crm_nlp_df_sorted_rev,cbs_resposne_bucket=self.cbs_resposne_bucket,years_list=self.years_list,year_excel_col_map_dict=year_excel_col_map_dict)


    def dump_cpl_data(self):
        bs_crm_nlp_df_sorted_rev = self.get_crm_nlp_df(statement_type="cpl")
        # month = get_month(fileid)
        year_excel_col_map_dict = self.get_years_excel_colmap(total_col=9)
        self.insert_records(report_workbook=self.workbook,bs_crm_nlp_df_sorted=bs_crm_nlp_df_sorted_rev,cbs_resposne_bucket=self.cbs_resposne_bucket,years_list=self.years_list,year_excel_col_map_dict=year_excel_col_map_dict)


    def dump_ccf_data(self):
        year_excel_col_map_dict = self.get_years_excel_colmap(self.years_list,total_col=8)
        # print(year_excel_col_map_dict)
        ccf_worksheet = self.workbook.get_sheet_by_name("CF")
        for idx,row in self.ccf_response_bucket.iterrows():
            if row["template_row_map"] is not None:
                if row["template_row_map"].isdigit():
                    # print(row)
                    # print(row["meta_keyword"])
                    for year in self.years_list:
                        year_column = year_excel_col_map_dict.get(int(year))
                        try:
                            ccf_worksheet.cell(row=int(row["template_row_map"]),column=year_column).value = row[str(year)]
                        except:
                            ccf_worksheet.cell(row=int(row["template_row_map"]),column=year_column).value = 0.0

    def save_excel(self):
        self.workbook.save(f"{self.fileid}.xlsx")

    # def insert_df(cdm_sheet_name,cdm_keyword_start_name,cdm_total_row_map,notes_table_df):
    #     pass

    # def get_month_year(self):
    #     pass # get this fro db